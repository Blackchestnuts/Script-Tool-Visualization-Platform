"""
测试效能平台（TEP）V1.3 - FastAPI 主应用
========================================
功能模块：
- Pytest 自动化测试执行引擎
- YAML / Excel 用例数据管理（模块化目录结构，支持双格式）
- 执行环境管理（CRUD，JSON 持久化）
- 通用脚本管理（Python / Shell / Bat 等）
- Allure 报告静态托管
- 任务状态管理与实时日志
"""

import os
import re
import json
import uuid
import shutil
import subprocess
import threading
from datetime import datetime
from pathlib import Path
from typing import Optional

import yaml
from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from jinja2 import Environment, FileSystemLoader
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import repositories as repo

# ============================================================
# 全局配置
# ============================================================
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
REPORTS_DIR = BASE_DIR / "reports"
TESTCASES_DIR = BASE_DIR / "testcases"
SCRIPTS_DIR = BASE_DIR / "scripts"
STATIC_DIR = BASE_DIR / "static"
TEMPLATES_DIR = BASE_DIR / "templates"
ENV_FILE = DATA_DIR / "environments.json"  # 环境配置持久化文件

# 确保目录存在
for d in [DATA_DIR, REPORTS_DIR, TESTCASES_DIR, SCRIPTS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ============================================================
# 脚本类型配置
# ============================================================
SCRIPT_TYPES = {
    ".py": {"label": "Python", "command": "python", "icon": "🐍", "color": "#3776ab"},
    ".sh": {"label": "Shell", "command": "bash", "icon": "🖥️", "color": "#4eaa25"},
    ".bat": {"label": "Batch", "command": "cmd", "icon": "📋", "color": "#0078d4"},
    ".ps1": {"label": "PowerShell", "command": "powershell", "icon": "⚡", "color": "#5391fe"},
    ".js": {"label": "Node.js", "command": "node", "icon": "🟢", "color": "#68a063"},
    ".rb": {"label": "Ruby", "command": "ruby", "icon": "💎", "color": "#cc342d"},
    ".lua": {"label": "Lua", "command": "lua", "icon": "🌙", "color": "#000080"},
}

# 用例文件支持的后缀集合
CASE_FILE_EXTENSIONS = {".yaml", ".yml", ".xlsx"}

# ============================================================
# 任务状态存储
# ============================================================
tasks_store: dict = {}
repo.ensure_database()

# ============================================================
# 环境配置管理（JSON 持久化）
# ============================================================

def load_environments() -> list:
    """从数据库加载环境配置列表，兼容首次启动时的 JSON 文件导入"""
    envs = repo.list_environments()
    if envs:
        return envs

    if ENV_FILE.exists():
        try:
            with open(ENV_FILE, "r", encoding="utf-8") as f:
                file_envs = json.load(f)
            save_environments(file_envs)
            return repo.list_environments()
        except (json.JSONDecodeError, Exception):
            pass

    default_envs = [
        {
            "id": "test_env",
            "name": "测试环境",
            "base_url": "http://test-api.example.com",
            "description": "日常测试环境",
            "db_host": "test-mysql.internal",
            "db_port": 3306,
            "timeout": 30,
        },
        {
            "id": "staging_env",
            "name": "预发环境",
            "base_url": "http://staging-api.example.com",
            "description": "预发布验证环境",
            "db_host": "staging-mysql.internal",
            "db_port": 3306,
            "timeout": 30,
        },
    ]
    save_environments(default_envs)
    return default_envs


def save_environments(envs: list):
    """保存环境配置到数据库，并保留 JSON 快照便于人工查看"""
    for env in envs:
        repo.upsert_environment(env)
    with open(ENV_FILE, "w", encoding="utf-8") as f:
        json.dump(envs, f, ensure_ascii=False, indent=2)


# ============================================================
# Pydantic 数据模型
# ============================================================

class RunRequest(BaseModel):
    """执行测试的请求体"""
    env: str = "test_env"
    markers: Optional[str] = None
    testcase_path: Optional[str] = None
    task_name: Optional[str] = None  # 任务名称，方便识别


class CaseUpdateRequest(BaseModel):
    """更新用例数据的请求体"""
    filename: str
    content: str
    module: Optional[str] = None  # 所属模块（子目录名）


class CaseCreateRequest(BaseModel):
    """创建用例的请求体"""
    filename: str
    module: str  # 所属模块
    content: str = ""
    file_type: str = "yaml"  # 文件类型: yaml / excel


class ModuleCreateRequest(BaseModel):
    """创建模块的请求体"""
    name: str
    description: str = ""


class ModuleRenameRequest(BaseModel):
    """重命名模块的请求体"""
    old_name: str
    new_name: str


class EnvCreateRequest(BaseModel):
    """创建环境的请求体"""
    id: str  # 环境唯一标识，如 test_env
    name: str  # 显示名称，如 测试环境
    base_url: str = ""
    description: str = ""
    db_host: str = ""
    db_port: int = 3306
    timeout: int = 30


class EnvUpdateRequest(BaseModel):
    """更新环境的请求体"""
    id: str
    name: Optional[str] = None
    base_url: Optional[str] = None
    description: Optional[str] = None
    db_host: Optional[str] = None
    db_port: Optional[int] = None
    timeout: Optional[int] = None


class ScriptUpdateRequest(BaseModel):
    filename: str
    content: str


class ScriptCreateRequest(BaseModel):
    filename: str
    content: str = ""
    description: str = ""


class ScriptRunRequest(BaseModel):
    filename: str
    args: Optional[str] = None
    task_name: Optional[str] = None  # 任务名称，方便识别


class TaskInfo(BaseModel):
    task_id: str
    task_name: Optional[str] = None  # 任务名称
    env: str
    status: str
    command: str
    created_at: str
    task_type: str = "pytest"
    finished_at: Optional[str] = None
    duration: Optional[str] = None
    exit_code: Optional[int] = None
    log: Optional[str] = None
    report_url: Optional[str] = None


# ============================================================
# FastAPI 应用实例
# ============================================================
app = FastAPI(
    title="测试效能平台 TEP V1.3",
    description="接口自动化测试 + 脚本管理 + 环境管理的 Web 管控平台（支持 YAML/Excel 双格式用例）",
    version="1.3.0",
)

# 挂载静态文件
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
if REPORTS_DIR.exists():
    app.mount("/reports", StaticFiles(directory=str(REPORTS_DIR), html=True), name="reports")

# Jinja2 模板引擎
jinja_env = Environment(loader=FileSystemLoader(str(TEMPLATES_DIR)), autoescape=True)


# ============================================================
# 工具函数
# ============================================================

def generate_task_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8]


def update_task_status(task_id: str, status: str, **kwargs):
    if task_id in tasks_store:
        tasks_store[task_id]["status"] = status
        for key, value in kwargs.items():
            tasks_store[task_id][key] = value
        if status in ("FINISHED", "ERROR"):
            tasks_store[task_id]["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            try:
                created = datetime.strptime(tasks_store[task_id]["created_at"], "%Y-%m-%d %H:%M:%S")
                finished = datetime.strptime(tasks_store[task_id]["finished_at"], "%Y-%m-%d %H:%M:%S")
                tasks_store[task_id]["duration"] = str(finished - created)
            except (ValueError, TypeError):
                pass
        repo.update_task(task_id, **tasks_store[task_id])


def save_task_log(task_id: str, stdout: str, stderr: str = ""):
    if task_id in tasks_store:
        tasks_store[task_id]["log"] = (stdout or "") + "\n" + (stderr or "")
        repo.save_task_log(task_id, tasks_store[task_id]["log"])


def get_script_type_info(suffix: str) -> dict:
    return SCRIPT_TYPES.get(suffix, {"label": "未知", "command": "", "icon": "📄", "color": "#999"})


def parse_script_info(filepath: Path) -> dict:
    suffix = filepath.suffix.lower()
    type_info = get_script_type_info(suffix)
    stat = filepath.stat()
    preview_lines = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            for i, line in enumerate(f):
                if i >= 5:
                    break
                preview_lines.append(line.rstrip())
    except Exception:
        pass
    return {
        "filename": filepath.name,
        "suffix": suffix,
        "type_label": type_info["label"],
        "type_icon": type_info["icon"],
        "type_color": type_info["color"],
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "preview": "\n".join(preview_lines),
    }


def get_module_dir(module_name: str) -> Path:
    """获取模块目录路径，并进行安全检查"""
    module_dir = DATA_DIR / module_name
    if not str(module_dir.resolve()).startswith(str(DATA_DIR.resolve())):
        raise ValueError("非法模块路径")
    return module_dir


def get_module_cases_files(module_dir: Path) -> list:
    """获取模块目录下所有用例文件（YAML + Excel）"""
    files = []
    for ext in CASE_FILE_EXTENSIONS:
        files.extend(module_dir.glob(f"*{ext}"))
    return sorted(files)


def get_case_file_type(filepath: Path) -> str:
    """根据后缀判断用例文件类型"""
    suffix = filepath.suffix.lower()
    if suffix == ".xlsx":
        return "excel"
    return "yaml"


# ============================================================
# Excel 解析与生成工具
# ============================================================

EXCEL_HEADERS = ["用例名称", "接口路径", "请求方法", "请求头", "请求参数", "预期状态码", "预期响应"]

EXCEL_HEADER_STYLE = {
    "font": Font(bold=True, color="FFFFFF", size=12),
    "fill": PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid"),
    "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
}

EXCEL_CELL_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

EXCEL_COL_WIDTHS = [25, 30, 12, 40, 40, 14, 40]

EXAMPLE_ROWS = [
    [
        "正常登录-成功",
        "/api/auth/login",
        "POST",
        '{"Content-Type": "application/json"}',
        '{"username": "admin", "password": "admin123"}',
        200,
        '{"code": 0, "message": "登录成功"}',
    ],
    [
        "密码错误-登录失败",
        "/api/auth/login",
        "POST",
        '{"Content-Type": "application/json"}',
        '{"username": "admin", "password": "wrong"}',
        200,
        '{"code": 1001, "message": "用户名或密码错误"}',
    ],
]


def _create_excel_template(filepath: Path, module_name: str = ""):
    """创建带标准表头的 Excel 用例模板文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 写入表头
    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = EXCEL_HEADER_STYLE["font"]
        cell.fill = EXCEL_HEADER_STYLE["fill"]
        cell.alignment = EXCEL_HEADER_STYLE["alignment"]
        cell.border = EXCEL_CELL_BORDER

    # 设置列宽
    for i, width in enumerate(EXCEL_COL_WIDTHS, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # 写入示例数据
    example_font = Font(color="6B7280", italic=True)
    for row_idx, row_data in enumerate(EXAMPLE_ROWS, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = example_font
            cell.border = EXCEL_CELL_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 添加说明 Sheet
    ws2 = wb.create_sheet(title="填写说明")
    instructions = [
        ["字段名称", "说明", "示例"],
        ["用例名称", "测试用例的名称，必填", "正常登录-成功"],
        ["接口路径", "API 接口路径，必填", "/api/auth/login"],
        ["请求方法", "HTTP 请求方法：GET/POST/PUT/DELETE/PATCH", "POST"],
        ["请求头", "JSON 格式的请求头，可选", '{"Content-Type": "application/json"}'],
        ["请求参数", "JSON 格式的请求参数，可选", '{"username": "admin"}'],
        ["预期状态码", "期望的 HTTP 响应状态码", "200"],
        ["预期响应", "JSON 格式的预期响应体，可选", '{"code": 0}'],
    ]
    for row in instructions:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 45

    wb.save(filepath)


def _write_cases_to_excel(filepath: Path, data: dict):
    """将结构化数据写入 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 写入表头
    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = EXCEL_HEADER_STYLE["font"]
        cell.fill = EXCEL_HEADER_STYLE["fill"]
        cell.alignment = EXCEL_HEADER_STYLE["alignment"]
        cell.border = EXCEL_CELL_BORDER

    # 设置列宽
    for i, width in enumerate(EXCEL_COL_WIDTHS, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    test_cases = data.get("test_cases", []) if isinstance(data, dict) else []
    for row_idx, case in enumerate(test_cases, 2):
        expected = case.get("expected", {}) if isinstance(case.get("expected"), dict) else {}
        row_data = [
            case.get("name", ""),
            case.get("endpoint", ""),
            case.get("method", "GET"),
            json.dumps(case.get("headers", {}), ensure_ascii=False) if case.get("headers") else "",
            json.dumps(case.get("params", {}), ensure_ascii=False) if case.get("params") else "",
            expected.get("status_code", ""),
            json.dumps(expected.get("body", {}), ensure_ascii=False) if expected.get("body") else "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = EXCEL_CELL_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(filepath)


def parse_excel_cases(filepath: Path, module_name: str = "") -> dict:
    """解析 Excel 用例文件，返回结构化数据

    Excel 格式要求：
    - 第一行为表头：用例名称, 接口路径, 请求方法, 请求头, 请求参数, 预期状态码, 预期响应
    - 从第二行起为用例数据
    """
    try:
        rel_path = str(filepath.relative_to(BASE_DIR))
    except ValueError:
        rel_path = str(filepath)

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return {
                "filename": filepath.name,
                "path": rel_path,
                "module": module_name,
                "file_type": "excel",
                "case_count": 0,
                "data": {"test_module": module_name, "test_cases": []},
            }

        # 读取表头
        header = [str(c or "").strip() for c in rows[0]]
        test_cases = []

        for row in rows[1:]:
            if not row or not row[0]:
                continue
            # 将行数据映射到表头
            row_dict = {}
            for i, h in enumerate(header):
                row_dict[h] = row[i] if i < len(row) else None

            case_item = {
                "name": str(row_dict.get("用例名称", "") or ""),
                "endpoint": str(row_dict.get("接口路径", "") or ""),
                "method": str(row_dict.get("请求方法", "GET") or "GET").upper(),
            }

            # 解析请求头（JSON 字符串）
            headers_str = row_dict.get("请求头", "")
            if headers_str:
                try:
                    case_item["headers"] = json.loads(str(headers_str))
                except (json.JSONDecodeError, TypeError):
                    case_item["headers"] = {"Content-Type": "application/json"}
            else:
                case_item["headers"] = {"Content-Type": "application/json"}

            # 解析请求参数（JSON 字符串）
            params_str = row_dict.get("请求参数", "")
            if params_str:
                try:
                    case_item["params"] = json.loads(str(params_str))
                except (json.JSONDecodeError, TypeError):
                    case_item["params"] = str(params_str)
            else:
                case_item["params"] = None

            # 解析预期结果
            expected = {}
            status_code = row_dict.get("预期状态码", "")
            if status_code:
                try:
                    expected["status_code"] = int(status_code)
                except (ValueError, TypeError):
                    expected["status_code"] = status_code

            expected_body = row_dict.get("预期响应", "")
            if expected_body:
                try:
                    expected["body"] = json.loads(str(expected_body))
                except (json.JSONDecodeError, TypeError):
                    expected["body"] = {"message": str(expected_body)}

            if expected:
                case_item["expected"] = expected

            test_cases.append(case_item)

        return {
            "filename": filepath.name,
            "path": rel_path,
            "module": module_name,
            "file_type": "excel",
            "case_count": len(test_cases),
            "data": {"test_module": module_name, "test_cases": test_cases},
        }
    except Exception as e:
        return {
            "filename": filepath.name,
            "path": rel_path,
            "module": module_name,
            "file_type": "excel",
            "case_count": 0,
            "error": f"Excel 解析错误: {str(e)}",
        }


def parse_yaml_cases(filepath: Path, module_name: str = "") -> dict:
    """解析单个 YAML 文件，返回结构化数据"""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
        try:
            rel_path = str(filepath.relative_to(BASE_DIR))
        except ValueError:
            rel_path = str(filepath)

        # 计算用例数量
        case_count = 0
        if data:
            if isinstance(data, list):
                case_count = len(data)
            elif isinstance(data, dict):
                if "test_cases" in data:
                    case_count = len(data["test_cases"])
                else:
                    case_count = 1

        return {
            "filename": filepath.name,
            "path": rel_path,
            "module": module_name,
            "file_type": "yaml",
            "case_count": case_count,
            "data": data,
        }
    except yaml.YAMLError as e:
        try:
            rel_path = str(filepath.relative_to(BASE_DIR))
        except ValueError:
            rel_path = str(filepath)
        return {
            "filename": filepath.name,
            "path": rel_path,
            "module": module_name,
            "file_type": "yaml",
            "case_count": 0,
            "error": f"YAML 解析错误: {str(e)}",
        }


def extract_summary_from_log(log: str) -> dict:
    summary = {"total": None, "passed": None, "failed": None, "errors": None, "skipped": None}
    if not log:
        return summary
    match = re.search(r"(\d+) passed", log)
    if match:
        summary["passed"] = int(match.group(1))
    match = re.search(r"(\d+) failed", log)
    if match:
        summary["failed"] = int(match.group(1))
    match = re.search(r"(\d+) error", log)
    if match:
        summary["errors"] = int(match.group(1))
    match = re.search(r"(\d+) skipped", log)
    if match:
        summary["skipped"] = int(match.group(1))
    if summary["passed"] is not None:
        total = summary["passed"]
        for k in ["failed", "errors", "skipped"]:
            if summary[k] is not None:
                total += summary[k]
        summary["total"] = total
    return summary


def _case_data_to_yaml(data: dict) -> str:
    return yaml.dump(data or {"test_cases": []}, allow_unicode=True, default_flow_style=False)


def bootstrap_database_from_files():
    """Import existing file-based data on first run so the DB is immediately usable."""
    load_environments()
    if repo.list_modules():
        return

    for module_dir in sorted(DATA_DIR.iterdir()):
        if not module_dir.is_dir() or module_dir.name.startswith("."):
            continue
        desc_file = module_dir / "_module.json"
        description = ""
        if desc_file.exists():
            try:
                with open(desc_file, "r", encoding="utf-8") as f:
                    description = json.load(f).get("description", "")
            except Exception:
                description = ""
        repo.upsert_module(module_dir.name, description)
        for filepath in get_module_cases_files(module_dir):
            if filepath.suffix.lower() == ".xlsx":
                parsed = parse_excel_cases(filepath, module_dir.name)
                file_type = "excel"
            else:
                parsed = parse_yaml_cases(filepath, module_dir.name)
                file_type = "yaml"
            data = parsed.get("data", {"test_module": module_dir.name, "test_cases": []})
            raw_content = parsed.get("raw_content") or _case_data_to_yaml(data)
            try:
                source_path = str(filepath.relative_to(BASE_DIR))
            except ValueError:
                source_path = str(filepath)
            repo.upsert_case_file(module_dir.name, filepath.name, file_type, data, raw_content, source_path)

    root_files = [f for f in DATA_DIR.glob("*.yaml")] + [f for f in DATA_DIR.glob("*.yml")]
    if root_files:
        root_module = "根目录"
        repo.upsert_module(root_module, "从 data 根目录迁移的历史用例")
        for filepath in sorted(root_files):
            parsed = parse_yaml_cases(filepath, root_module)
            data = parsed.get("data", {"test_module": root_module, "test_cases": []})
            repo.upsert_case_file(
                root_module,
                filepath.name,
                "yaml",
                data,
                parsed.get("raw_content") or _case_data_to_yaml(data),
                str(filepath.relative_to(BASE_DIR)),
            )


bootstrap_database_from_files()


# ============================================================
# 执行引擎
# ============================================================

def run_pytest_task(task_id: str, env: str, markers: Optional[str] = None, testcase_path: Optional[str] = None):
    update_task_status(task_id, "RUNNING")
    test_path = testcase_path or str(TESTCASES_DIR)

    command_parts = [
        "python", "-m", "pytest", test_path,
        f"--env={env}",
        f"--alluredir={REPORTS_DIR / task_id}",
        "-v", "--tb=short",
    ]
    if markers:
        command_parts.extend(["-m", markers])

    tasks_store[task_id]["command"] = " ".join(command_parts)
    repo.update_task(task_id, command=tasks_store[task_id]["command"])

    try:
        result = subprocess.run(
            command_parts, capture_output=True, text=True, timeout=600,
            cwd=str(BASE_DIR), env={**os.environ, "TEST_ENV": env},
        )
        save_task_log(task_id, result.stdout, result.stderr)
        tasks_store[task_id]["exit_code"] = result.returncode
        update_task_status(task_id, "FINISHED")

        report_dir = REPORTS_DIR / task_id
        html_report_dir = report_dir / "html"
        html_report_dir.mkdir(parents=True, exist_ok=True)
        try:
            subprocess.run(
                ["allure", "generate", str(report_dir), "-o", str(html_report_dir), "--clean"],
                capture_output=True, text=True, timeout=120,
            )
            tasks_store[task_id]["report_url"] = f"/reports/{task_id}/html/index.html"
            repo.upsert_report(task_id, tasks_store[task_id]["report_url"], str(html_report_dir), True)
        except Exception:
            tasks_store[task_id]["report_url"] = f"/reports/{task_id}/"
            repo.upsert_report(task_id, tasks_store[task_id]["report_url"], str(report_dir), False)

    except subprocess.TimeoutExpired:
        update_task_status(task_id, "ERROR")
        save_task_log(task_id, "", "测试执行超时（超过600秒）")
        tasks_store[task_id]["exit_code"] = -1
    except Exception as e:
        update_task_status(task_id, "ERROR")
        save_task_log(task_id, "", str(e))
        tasks_store[task_id]["exit_code"] = -1


def run_script_task(task_id: str, filename: str, args: Optional[str] = None):
    update_task_status(task_id, "RUNNING")
    filepath = SCRIPTS_DIR / filename
    suffix = filepath.suffix.lower()
    type_info = get_script_type_info(suffix)

    if suffix == ".bat":
        command_parts = ["cmd", "/c", str(filepath)]
    elif suffix == ".ps1":
        command_parts = ["powershell", "-ExecutionPolicy", "Bypass", "-File", str(filepath)]
    else:
        command_parts = [type_info["command"], str(filepath)]

    if args:
        command_parts.extend(args.split())

    tasks_store[task_id]["command"] = " ".join(command_parts)
    repo.update_task(task_id, command=tasks_store[task_id]["command"])

    try:
        result = subprocess.run(
            command_parts, capture_output=True, text=True, timeout=600,
            cwd=str(BASE_DIR), env={**os.environ, "SCRIPT_NAME": filename},
        )
        save_task_log(task_id, result.stdout, result.stderr)
        tasks_store[task_id]["exit_code"] = result.returncode
        update_task_status(task_id, "FINISHED")
    except subprocess.TimeoutExpired:
        update_task_status(task_id, "ERROR")
        save_task_log(task_id, "", "脚本执行超时")
        tasks_store[task_id]["exit_code"] = -1
    except FileNotFoundError:
        update_task_status(task_id, "ERROR")
        save_task_log(task_id, "", f"找不到执行器: {type_info['command']}")
        tasks_store[task_id]["exit_code"] = -1
    except Exception as e:
        update_task_status(task_id, "ERROR")
        save_task_log(task_id, "", str(e))
        tasks_store[task_id]["exit_code"] = -1


# ============================================================
# 页面路由
# ============================================================

def render_template(template_name: str, **context) -> HTMLResponse:
    template = jinja_env.get_template(template_name)
    html = template.render(**context)
    return HTMLResponse(html)


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return render_template("index.html", request=request)


@app.get("/cases", response_class=HTMLResponse)
async def cases_page(request: Request):
    return render_template("cases.html", request=request)


@app.get("/scripts", response_class=HTMLResponse)
async def scripts_page(request: Request):
    return render_template("scripts.html", request=request)


@app.get("/reports-page", response_class=HTMLResponse)
async def reports_page(request: Request):
    return render_template("reports.html", request=request)


# ============================================================
# API 路由 - 环境管理
# ============================================================

@app.get("/api/environments")
async def get_environments():
    """获取所有执行环境"""
    envs = load_environments()
    return {"environments": envs, "total": len(envs)}


@app.post("/api/environments")
async def create_environment(request: EnvCreateRequest):
    """新增执行环境"""
    new_env = {
        "id": request.id,
        "name": request.name,
        "base_url": request.base_url,
        "description": request.description,
        "db_host": request.db_host,
        "db_port": request.db_port,
        "timeout": request.timeout,
    }
    try:
        new_env = repo.create_environment(new_env)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"环境标识 '{request.id}' 已存在")

    # 同步更新 config.yaml 供 pytest 使用
    envs = load_environments()
    _sync_config_yaml(envs)

    return {"message": f"环境 '{request.name}' 创建成功", "environment": new_env}


@app.put("/api/environments")
async def update_environment(request: EnvUpdateRequest):
    """更新执行环境"""
    target = repo.update_environment(
        request.id,
        {
            "name": request.name,
            "base_url": request.base_url,
            "description": request.description,
            "db_host": request.db_host,
            "db_port": request.db_port,
            "timeout": request.timeout,
        },
    )
    if not target:
        raise HTTPException(status_code=404, detail=f"环境 '{request.id}' 不存在")

    envs = load_environments()
    _sync_config_yaml(envs)

    return {"message": f"环境 '{request.id}' 更新成功", "environment": target}


@app.delete("/api/environments/{env_id}")
async def delete_environment(env_id: str):
    """删除执行环境"""
    envs = load_environments()

    # 至少保留一个环境
    if len(envs) <= 1:
        raise HTTPException(status_code=400, detail="至少需要保留一个执行环境")

    if not repo.delete_environment(env_id):
        raise HTTPException(status_code=404, detail=f"环境 '{env_id}' 不存在")

    new_envs = load_environments()
    _sync_config_yaml(new_envs)

    return {"message": f"环境 '{env_id}' 删除成功", "total": len(new_envs)}


def _sync_config_yaml(envs: list):
    """将环境配置同步写入 config.yaml，供 pytest conftest 读取"""
    config = {}
    for e in envs:
        config[e["id"]] = {
            "base_url": e.get("base_url", ""),
            "db_host": e.get("db_host", ""),
            "db_port": e.get("db_port", 3306),
            "timeout": e.get("timeout", 30),
        }
    config_path = BASE_DIR / "config.yaml"
    with open(config_path, "w", encoding="utf-8") as f:
        yaml.dump(config, f, allow_unicode=True, default_flow_style=False)


# ============================================================
# API 路由 - 模块管理
# ============================================================

@app.get("/api/modules")
async def get_modules():
    """获取所有用例模块"""
    modules = repo.list_modules()
    return {"modules": modules, "total": len(modules)}


@app.post("/api/modules")
async def create_module(request: ModuleCreateRequest):
    """创建新模块（在 data/ 下新建子目录）"""
    module_name = request.name.strip()

    # 校验名称合法性
    if not module_name or not re.match(r'^[\w\u4e00-\u9fff]+$', module_name):
        raise HTTPException(status_code=400, detail="模块名称只能包含中文、字母、数字和下划线")

    if repo.get_module(module_name):
        raise HTTPException(status_code=400, detail=f"模块 '{module_name}' 已存在")

    module_dir = get_module_dir(module_name)
    module_dir.mkdir(parents=True, exist_ok=True)

    # 写入模块描述文件
    if request.description:
        desc_file = module_dir / "_module.json"
        with open(desc_file, "w", encoding="utf-8") as f:
            json.dump({"name": module_name, "description": request.description}, f, ensure_ascii=False, indent=2)

    repo.create_module(module_name, request.description)
    return {"message": f"模块 '{module_name}' 创建成功", "name": module_name}


@app.put("/api/modules")
async def rename_module(request: ModuleRenameRequest):
    """重命名模块"""
    if not request.new_name or not re.match(r'^[\w\u4e00-\u9fff]+$', request.new_name):
        raise HTTPException(status_code=400, detail="非法模块名称")

    if not repo.get_module(request.old_name):
        raise HTTPException(status_code=404, detail=f"模块 '{request.old_name}' 不存在")
    if repo.get_module(request.new_name):
        raise HTTPException(status_code=400, detail=f"模块 '{request.new_name}' 已存在")

    try:
        old_dir = get_module_dir(request.old_name)
        new_dir = get_module_dir(request.new_name)
        if old_dir.exists() and not new_dir.exists():
            old_dir.rename(new_dir)
    except ValueError:
        raise HTTPException(status_code=400, detail="非法模块名称")

    try:
        repo.rename_module(request.old_name, request.new_name)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"模块 '{request.new_name}' 已存在")
    return {"message": f"模块 '{request.old_name}' 已重命名为 '{request.new_name}'"}


@app.delete("/api/modules/{module_name}")
async def delete_module(module_name: str):
    """删除模块及其所有用例"""
    if not repo.get_module(module_name):
        raise HTTPException(status_code=404, detail=f"模块 '{module_name}' 不存在")

    if not repo.delete_module(module_name):
        raise HTTPException(status_code=404, detail=f"模块 '{module_name}' 不存在")

    try:
        module_dir = get_module_dir(module_name)
        if module_dir.exists():
            shutil.rmtree(module_dir)
    except ValueError:
        pass
    return {"message": f"模块 '{module_name}' 及其所有用例已删除"}


# ============================================================
# API 路由 - 用例管理（模块化，支持 YAML/Excel 双格式）
# ============================================================

@app.get("/api/cases")
async def get_cases(module: Optional[str] = None):
    """获取用例列表，支持按模块过滤，同时包含 YAML 和 Excel 导入记录"""
    if module:
        if not repo.get_module(module):
            raise HTTPException(status_code=404, detail=f"模块 '{module}' 不存在")
        cases = repo.list_case_files(module)
        return {"cases": cases, "total": len(cases), "module": module}
    else:
        all_cases = repo.list_case_files()
        return {"cases": all_cases, "total": len(all_cases)}


@app.get("/api/cases/{module}/{filename}")
async def get_case_detail(module: str, filename: str):
    """获取单个用例文件详情"""
    case_data = repo.get_case_file(module, filename)
    if not case_data:
        raise HTTPException(status_code=404, detail=f"文件 {module}/{filename} 不存在")

    if not case_data.get("raw_content"):
        case_data["raw_content"] = _case_data_to_yaml(case_data.get("data", {}))
    return case_data


@app.post("/api/cases/create")
async def create_case(request: CaseCreateRequest):
    """在指定模块下创建用例文件（支持 YAML 和 Excel 格式）"""
    if not repo.get_module(request.module):
        raise HTTPException(status_code=404, detail=f"模块 '{request.module}' 不存在")

    module_dir = get_module_dir(request.module)
    module_dir.mkdir(parents=True, exist_ok=True)
    file_type = request.file_type.lower()
    filename = request.filename

    if file_type == "excel":
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        filepath = module_dir / filename
        if repo.get_case_file(request.module, filename):
            raise HTTPException(status_code=400, detail=f"文件 {filename} 已存在")
        if not str(filepath.resolve()).startswith(str(module_dir.resolve())):
            raise HTTPException(status_code=400, detail="非法文件路径")

        _create_excel_template(filepath, request.module)
        data = {"test_module": request.module, "test_cases": []}
        repo.create_case_file(
            request.module,
            filename,
            "excel",
            data,
            raw_content=_case_data_to_yaml(data),
            source_path=str(filepath.relative_to(BASE_DIR)),
        )

        return {"message": f"Excel 文件 {request.module}/{filename} 创建成功", "filename": filename, "module": request.module, "file_type": "excel"}
    else:
        if not filename.endswith((".yaml", ".yml")):
            filename += ".yaml"

        filepath = module_dir / filename
        if repo.get_case_file(request.module, filename):
            raise HTTPException(status_code=400, detail=f"文件 {filename} 已存在")
        if not str(filepath.resolve()).startswith(str(module_dir.resolve())):
            raise HTTPException(status_code=400, detail="非法文件路径")

        content = request.content or f"# {request.module} - {filename}\ntest_module: {request.module}\ntest_cases: []\n"
        if content:
            try:
                data = yaml.safe_load(content) or {"test_module": request.module, "test_cases": []}
            except yaml.YAMLError as e:
                raise HTTPException(status_code=400, detail=f"YAML 格式错误: {str(e)}")

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)
        repo.create_case_file(
            request.module,
            filename,
            "yaml",
            data,
            raw_content=content,
            source_path=str(filepath.relative_to(BASE_DIR)),
        )

        return {"message": f"YAML 文件 {request.module}/{filename} 创建成功", "filename": filename, "module": request.module, "file_type": "yaml"}


@app.post("/api/cases/update")
async def update_case(request: CaseUpdateRequest):
    """更新用例数据（支持 YAML 和 Excel 格式）"""
    module = request.module or ""
    if not module:
        raise HTTPException(status_code=400, detail="请指定模块名称")
    case_info = repo.get_case_file(module, request.filename)
    if not case_info:
        raise HTTPException(status_code=404, detail=f"文件 {request.filename} 不存在")

    try:
        data = yaml.safe_load(request.content) or {"test_module": module, "test_cases": []}
    except yaml.YAMLError as e:
        raise HTTPException(status_code=400, detail=f"YAML 格式错误: {str(e)}")

    repo.update_case_file(module, request.filename, data, raw_content=request.content)

    try:
        base = get_module_dir(module)
        base.mkdir(parents=True, exist_ok=True)
        filepath = base / request.filename
        if not str(filepath.resolve()).startswith(str(base.resolve())):
            raise HTTPException(status_code=400, detail="非法文件路径")
        if case_info.get("file_type") == "excel":
            _write_cases_to_excel(filepath, data)
        else:
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(request.content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"写入备份文件失败: {str(e)}")

    return {"message": f"文件 {request.filename} 更新成功"}


@app.delete("/api/cases/{module}/{filename}")
async def delete_case(module: str, filename: str):
    """删除用例文件"""
    if not repo.delete_case_file(module, filename):
        raise HTTPException(status_code=404, detail=f"文件 {module}/{filename} 不存在")

    try:
        module_dir = get_module_dir(module)
        filepath = module_dir / filename
        if filepath.exists() and str(filepath.resolve()).startswith(str(module_dir.resolve())):
            filepath.unlink()
    except ValueError:
        pass
    return {"message": f"文件 {module}/{filename} 删除成功"}


# ============================================================
# API 路由 - Excel 用例导入 / 导出 / 模板下载
# ============================================================

@app.get("/api/cases/excel-template")
async def download_excel_template():
    """下载 Excel 用例导入模板"""
    template_dir = BASE_DIR / "temp"
    template_dir.mkdir(exist_ok=True)
    template_path = template_dir / "用例导入模板.xlsx"
    _create_excel_template(template_path)
    return FileResponse(
        path=str(template_path),
        filename="用例导入模板.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/cases/import-excel")
async def import_excel_cases(
    module: str = "",
    file: UploadFile = File(...),
):
    """导入 Excel 用例文件到指定模块

    - 支持上传 .xlsx 文件
    - 自动解析 Excel 内容并保存到模块目录
    - 同时生成对应的 YAML 版本
    """
    if not module:
        raise HTTPException(status_code=400, detail="请指定导入的目标模块")

    if not repo.get_module(module):
        raise HTTPException(status_code=404, detail=f"模块 '{module}' 不存在")
    module_dir = get_module_dir(module)
    module_dir.mkdir(parents=True, exist_ok=True)

    # 检查文件类型
    suffix = Path(file.filename).suffix.lower()
    if suffix != ".xlsx":
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式的 Excel 文件")

    # 保存 Excel 文件
    import_filename = file.filename
    filepath = module_dir / import_filename
    if not str(filepath.resolve()).startswith(str(module_dir.resolve())):
        raise HTTPException(status_code=400, detail="非法文件路径")

    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    # 解析 Excel 内容
    case_data = parse_excel_cases(filepath, module)
    data = case_data.get("data", {"test_module": module, "test_cases": []})
    repo.upsert_case_file(
        module,
        import_filename,
        "excel",
        data,
        raw_content=_case_data_to_yaml(data),
        source_path=str(filepath.relative_to(BASE_DIR)),
    )

    # 同时生成 YAML 版本（方便用户切换格式）
    yaml_filename = Path(import_filename).stem + ".yaml"
    yaml_filepath = module_dir / yaml_filename
    if case_data.get("data") and not repo.get_case_file(module, yaml_filename):
        yaml_content = yaml.dump(data, allow_unicode=True, default_flow_style=False)
        with open(yaml_filepath, "w", encoding="utf-8") as f:
            f.write(yaml_content)
        repo.create_case_file(
            module,
            yaml_filename,
            "yaml",
            data,
            raw_content=yaml_content,
            source_path=str(yaml_filepath.relative_to(BASE_DIR)),
        )

    return {
        "message": f"Excel 文件导入成功: {module}/{import_filename}",
        "filename": import_filename,
        "module": module,
        "file_type": "excel",
        "case_count": case_data.get("case_count", 0),
        "yaml_generated": yaml_filename if case_data.get("data") else None,
    }


@app.post("/api/cases/export-yaml/{module}/{filename}")
async def export_yaml_from_excel(module: str, filename: str):
    """将 Excel 用例导出为 YAML 格式"""
    case_data = repo.get_case_file(module, filename)
    if not case_data:
        raise HTTPException(status_code=404, detail=f"文件 {module}/{filename} 不存在")

    if case_data.get("file_type") != "excel":
        raise HTTPException(status_code=400, detail="仅支持将 Excel 文件导出为 YAML")

    yaml_filename = Path(filename).stem + ".yaml"
    if repo.get_case_file(module, yaml_filename):
        raise HTTPException(status_code=400, detail=f"YAML 文件 {yaml_filename} 已存在，请先删除")

    module_dir = get_module_dir(module)
    module_dir.mkdir(parents=True, exist_ok=True)
    yaml_filepath = module_dir / yaml_filename

    if case_data.get("data"):
        yaml_content = yaml.dump(case_data["data"], allow_unicode=True, default_flow_style=False)
        with open(yaml_filepath, "w", encoding="utf-8") as f:
            f.write(yaml_content)
        repo.create_case_file(
            module,
            yaml_filename,
            "yaml",
            case_data["data"],
            raw_content=yaml_content,
            source_path=str(yaml_filepath.relative_to(BASE_DIR)),
        )
        return {"message": f"已导出 YAML 文件: {module}/{yaml_filename}", "yaml_filename": yaml_filename}
    else:
        raise HTTPException(status_code=400, detail="Excel 文件中无有效用例数据")


@app.post("/api/cases/export-excel/{module}/{filename}")
async def export_excel_from_yaml(module: str, filename: str):
    """将 YAML 用例导出为 Excel 格式"""
    case_data = repo.get_case_file(module, filename)
    if not case_data:
        raise HTTPException(status_code=404, detail=f"文件 {module}/{filename} 不存在")

    if case_data.get("file_type") != "yaml":
        raise HTTPException(status_code=400, detail="仅支持将 YAML 文件导出为 Excel")

    excel_filename = Path(filename).stem + ".xlsx"
    if repo.get_case_file(module, excel_filename):
        raise HTTPException(status_code=400, detail=f"Excel 文件 {excel_filename} 已存在，请先删除")

    module_dir = get_module_dir(module)
    module_dir.mkdir(parents=True, exist_ok=True)
    excel_filepath = module_dir / excel_filename

    if case_data.get("data"):
        _write_cases_to_excel(excel_filepath, case_data["data"])
        repo.create_case_file(
            module,
            excel_filename,
            "excel",
            case_data["data"],
            raw_content=_case_data_to_yaml(case_data["data"]),
            source_path=str(excel_filepath.relative_to(BASE_DIR)),
        )
        return {"message": f"已导出 Excel 文件: {module}/{excel_filename}", "excel_filename": excel_filename}
    else:
        raise HTTPException(status_code=400, detail="YAML 文件中无有效用例数据")


# ============================================================
# API 路由 - 脚本管理
# ============================================================

@app.get("/api/scripts")
async def get_scripts():
    supported_suffixes = set(SCRIPT_TYPES.keys())
    script_files = [f for f in SCRIPTS_DIR.iterdir() if f.is_file() and f.suffix.lower() in supported_suffixes]
    if not script_files:
        return {"scripts": [], "total": 0}
    scripts = []
    for f in sorted(script_files):
        info = parse_script_info(f)
        repo.upsert_script(f.name, f.suffix.lower(), info["type_label"], f)
        scripts.append(info)
    return {"scripts": scripts, "total": len(scripts)}


@app.get("/api/scripts/{filename}")
async def get_script_detail(filename: str):
    filepath = SCRIPTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail=f"脚本 {filename} 不存在")
    if not str(filepath.resolve()).startswith(str(SCRIPTS_DIR.resolve())):
        raise HTTPException(status_code=400, detail="非法文件路径")
    info = parse_script_info(filepath)
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        info["content"] = f.read()
    return info


@app.post("/api/scripts/create")
async def create_script(request: ScriptCreateRequest):
    suffix = Path(request.filename).suffix.lower()
    if suffix not in SCRIPT_TYPES:
        raise HTTPException(status_code=400, detail=f"不支持的脚本类型: {suffix}")
    filepath = SCRIPTS_DIR / request.filename
    if filepath.exists():
        raise HTTPException(status_code=400, detail=f"脚本 {request.filename} 已存在")
    if not str(filepath.resolve()).startswith(str(SCRIPTS_DIR.resolve())):
        raise HTTPException(status_code=400, detail="非法文件路径")
    content = request.content or f"# {request.description or request.filename}\n"
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)
    type_info = get_script_type_info(suffix)
    repo.upsert_script(request.filename, suffix, type_info["label"], filepath, request.description)
    return {"message": f"脚本 {request.filename} 创建成功", "filename": request.filename}


@app.post("/api/scripts/update")
async def update_script(request: ScriptUpdateRequest):
    filepath = SCRIPTS_DIR / request.filename
    if not str(filepath.resolve()).startswith(str(SCRIPTS_DIR.resolve())):
        raise HTTPException(status_code=400, detail="非法文件路径")
    if not filepath.exists():
        raise HTTPException(status_code=404, detail=f"脚本 {request.filename} 不存在")
    try:
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(request.content)
        suffix = filepath.suffix.lower()
        type_info = get_script_type_info(suffix)
        repo.upsert_script(request.filename, suffix, type_info["label"], filepath)
        return {"message": f"脚本 {request.filename} 更新成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/scripts/{filename}")
async def delete_script(filename: str):
    filepath = SCRIPTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail=f"脚本 {filename} 不存在")
    if not str(filepath.resolve()).startswith(str(SCRIPTS_DIR.resolve())):
        raise HTTPException(status_code=400, detail="非法文件路径")
    filepath.unlink()
    repo.delete_script_record(filename)
    return {"message": f"脚本 {filename} 删除成功"}


@app.post("/api/scripts/upload")
async def upload_script(file: UploadFile = File(...)):
    suffix = Path(file.filename).suffix.lower()
    if suffix not in SCRIPT_TYPES:
        raise HTTPException(status_code=400, detail=f"不支持的脚本类型: {suffix}")
    filepath = SCRIPTS_DIR / file.filename
    if not str(filepath.resolve()).startswith(str(SCRIPTS_DIR.resolve())):
        raise HTTPException(status_code=400, detail="非法文件路径")
    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)
    type_info = get_script_type_info(suffix)
    repo.upsert_script(file.filename, suffix, type_info["label"], filepath)
    return {"message": f"脚本 {file.filename} 上传成功", "filename": file.filename}


@app.post("/api/scripts/run")
async def run_script(request: ScriptRunRequest):
    filepath = SCRIPTS_DIR / request.filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail=f"脚本 {request.filename} 不存在")
    suffix = filepath.suffix.lower()
    if suffix not in SCRIPT_TYPES:
        raise HTTPException(status_code=400, detail=f"不支持的脚本类型: {suffix}")

    # 自动生成默认任务名称
    task_name = request.task_name or f"运行 {request.filename}"

    task_id = generate_task_id()
    tasks_store[task_id] = {
        "task_id": task_id, "task_name": task_name, "env": "script", "status": "PENDING", "command": "",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "task_type": "script", "script_name": request.filename,
        "finished_at": None, "duration": None, "exit_code": None, "log": None, "report_url": None,
    }
    repo.create_task(tasks_store[task_id])
    thread = threading.Thread(target=run_script_task, args=(task_id, request.filename, request.args), daemon=True)
    thread.start()
    return {"task_id": task_id, "task_name": task_name, "status": "PENDING", "message": f"脚本 {request.filename} 已提交执行"}


# ============================================================
# API 路由 - 任务执行
# ============================================================

@app.post("/api/run")
async def run_tests(request: RunRequest):
    """触发 Pytest 测试执行"""
    # 动态验证环境（从 environments.json 读取）
    envs = load_environments()
    valid_env_ids = [e["id"] for e in envs]
    if request.env not in valid_env_ids:
        raise HTTPException(status_code=400, detail=f"无效的环境: {request.env}，可选: {valid_env_ids}")

    # 自动生成默认任务名称
    env_info = next((e for e in envs if e["id"] == request.env), None)
    env_name = env_info["name"] if env_info else request.env
    task_name = request.task_name or f"{env_name} 接口测试"

    task_id = generate_task_id()
    tasks_store[task_id] = {
        "task_id": task_id, "task_name": task_name, "env": request.env, "status": "PENDING", "command": "",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "task_type": "pytest", "script_name": None,
        "finished_at": None, "duration": None, "exit_code": None, "log": None, "report_url": None,
    }
    repo.create_task(tasks_store[task_id])
    report_dir = REPORTS_DIR / task_id
    report_dir.mkdir(parents=True, exist_ok=True)

    thread = threading.Thread(
        target=run_pytest_task,
        args=(task_id, request.env, request.markers, request.testcase_path),
        daemon=True,
    )
    thread.start()
    return {"task_id": task_id, "task_name": task_name, "status": "PENDING", "message": "测试任务已提交"}


@app.get("/api/status/{task_id}")
async def get_task_status(task_id: str):
    task = tasks_store.get(task_id) or repo.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"任务 {task_id} 不存在")
    summary = extract_summary_from_log(task.get("log", ""))
    return {**task, "summary": summary}


@app.get("/api/tasks")
async def get_tasks(limit: int = 20, status: Optional[str] = None, task_type: Optional[str] = None):
    tasks = repo.list_tasks(limit=limit, status=status, task_type=task_type)
    for task in tasks:
        task["summary"] = extract_summary_from_log(task.get("log", ""))
    return {"tasks": tasks, "total": len(tasks)}


# ============================================================
# API 路由 - 报告管理
# ============================================================

@app.get("/api/reports")
async def get_reports():
    if not REPORTS_DIR.exists():
        return {"reports": [], "total": 0}
    reports_by_id = {r["task_id"]: r for r in repo.list_reports()}
    reports = []
    for d in sorted(REPORTS_DIR.iterdir(), reverse=True):
        if d.is_dir() and not d.name.startswith("."):
            html_dir = d / "html"
            has_html = html_dir.exists() and (html_dir / "index.html").exists()
            task_info = tasks_store.get(d.name) or repo.get_task(d.name) or reports_by_id.get(d.name, {})
            report_entry = {
                "task_id": d.name,
                "created_at": task_info.get("created_at", ""),
                "env": task_info.get("env", ""),
                "has_html_report": has_html,
                "report_url": f"/reports/{d.name}/html/index.html" if has_html else f"/reports/{d.name}/",
                "status": task_info.get("status", "UNKNOWN"),
                "task_type": task_info.get("task_type", "unknown"),
                "script_name": task_info.get("script_name", ""),
            }
            if task_info:
                report_entry["summary"] = extract_summary_from_log(task_info.get("log", ""))
            else:
                report_entry["summary"] = {}
            repo.upsert_report(d.name, report_entry["report_url"], str(html_dir if has_html else d), has_html)
            reports.append(report_entry)
    known_ids = {r["task_id"] for r in reports}
    for report_entry in reports_by_id.values():
        if report_entry["task_id"] in known_ids:
            continue
        report_entry["summary"] = extract_summary_from_log(report_entry.pop("log", ""))
        reports.append(report_entry)
    return {"reports": reports, "total": len(reports)}


@app.get("/api/log/{task_id}")
async def get_task_log(task_id: str, tail: Optional[int] = None):
    task = tasks_store.get(task_id) or repo.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"任务 {task_id} 不存在")
    log = task.get("log", "")
    if tail and log:
        log = "\n".join(log.splitlines()[-tail:])
    return {"task_id": task_id, "log": log}


# ============================================================
# 健康检查 & 类型查询
# ============================================================

@app.get("/api/health")
async def health_check():
    return {
        "status": "ok",
        "version": "1.3.0",
        "running_tasks": len([t for t in tasks_store.values() if t["status"] == "RUNNING"]),
        "total_tasks": len(tasks_store),
    }


@app.get("/api/script-types")
async def get_script_types():
    result = []
    for suffix, info in SCRIPT_TYPES.items():
        result.append({"suffix": suffix, "label": info["label"], "command": info["command"],
                       "icon": info["icon"], "color": info["color"]})
    return {"types": result}


# ============================================================
# 启动入口
# ============================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True, log_level="info")
