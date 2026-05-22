"""
Migrate existing TEP file data into data/tep.db.

The script is idempotent: environments, modules and case files are upserted.
Original JSON/YAML/Excel files are left untouched so they can still be used as
backups or import/export artifacts.
"""
import json
import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

import repositories as repo  # noqa: E402


DATA_DIR = ROOT_DIR / "data"
CASE_FILE_EXTENSIONS = {".yaml", ".yml", ".xlsx"}
EXCEL_HEADERS = ["用例名称", "接口路径", "请求方法", "请求头", "请求参数", "预期状态码", "预期响应"]


def default_environments() -> list:
    return [
        {
            "id": "test_env",
            "name": "测试环境",
            "base_url": "http://test-api.example.com",
            "description": "日常测试环境",
            "db_host": "test-mysql.internal",
            "db_port": 3306,
            "timeout": 30,
        },
        {
            "id": "staging_env",
            "name": "预发环境",
            "base_url": "http://staging-api.example.com",
            "description": "预发布验证环境",
            "db_host": "staging-mysql.internal",
            "db_port": 3306,
            "timeout": 30,
        },
    ]


def parse_yaml_cases(filepath: Path, module_name: str) -> tuple[dict, str]:
    raw = filepath.read_text(encoding="utf-8")
    data = yaml.safe_load(raw) or {"test_module": module_name, "test_cases": []}
    if isinstance(data, list):
        data = {"test_module": module_name, "test_cases": data}
    if "test_cases" not in data:
        data = {"test_module": module_name, "test_cases": [data]}
    return data, raw


def parse_json_cell(value, fallback=None):
    if value in (None, ""):
        return fallback
    try:
        return json.loads(str(value))
    except (json.JSONDecodeError, TypeError):
        return fallback if fallback is not None else str(value)


def parse_excel_cases(filepath: Path, module_name: str) -> tuple[dict, str]:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return {"test_module": module_name, "test_cases": []}, ""

    header = [str(c or "").strip() for c in rows[0]]
    cases = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        row_dict = {h: row[i] if i < len(row) else None for i, h in enumerate(header)}
        expected = {}
        status_code = row_dict.get("预期状态码")
        if status_code not in (None, ""):
            try:
                expected["status_code"] = int(status_code)
            except (ValueError, TypeError):
                expected["status_code"] = status_code
        expected_body = parse_json_cell(row_dict.get("预期响应"), None)
        if expected_body is not None:
            expected["body"] = expected_body
        cases.append(
            {
                "name": str(row_dict.get("用例名称", "") or ""),
                "endpoint": str(row_dict.get("接口路径", "") or ""),
                "method": str(row_dict.get("请求方法", "GET") or "GET").upper(),
                "headers": parse_json_cell(row_dict.get("请求头"), {"Content-Type": "application/json"}),
                "params": parse_json_cell(row_dict.get("请求参数"), None),
                "expected": expected,
            }
        )
    data = {"test_module": module_name, "test_cases": cases}
    return data, yaml.dump(data, allow_unicode=True, default_flow_style=False)


def migrate_environments():
    env_file = DATA_DIR / "environments.json"
    if env_file.exists():
        try:
            envs = json.loads(env_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            envs = default_environments()
    else:
        envs = default_environments()
    for env in envs:
        repo.upsert_environment(env)
    return len(envs)


def migrate_modules_and_cases():
    migrated_files = 0
    for module_dir in sorted(DATA_DIR.iterdir() if DATA_DIR.exists() else []):
        if not module_dir.is_dir() or module_dir.name.startswith("."):
            continue
        desc_file = module_dir / "_module.json"
        description = ""
        if desc_file.exists():
            try:
                description = json.loads(desc_file.read_text(encoding="utf-8")).get("description", "")
            except json.JSONDecodeError:
                description = ""
        repo.upsert_module(module_dir.name, description)
        for filepath in sorted(module_dir.iterdir()):
            if filepath.suffix.lower() not in CASE_FILE_EXTENSIONS:
                continue
            if filepath.suffix.lower() == ".xlsx":
                data, raw = parse_excel_cases(filepath, module_dir.name)
                file_type = "excel"
            else:
                data, raw = parse_yaml_cases(filepath, module_dir.name)
                file_type = "yaml"
            repo.upsert_case_file(
                module_dir.name,
                filepath.name,
                file_type,
                data,
                raw_content=raw,
                source_path=str(filepath.relative_to(ROOT_DIR)),
            )
            migrated_files += 1

    root_module = "根目录"
    root_files = [p for p in DATA_DIR.glob("*.yaml")] + [p for p in DATA_DIR.glob("*.yml")]
    if root_files:
        repo.upsert_module(root_module, "从 data 根目录迁移的历史用例")
        for filepath in sorted(root_files):
            data, raw = parse_yaml_cases(filepath, root_module)
            repo.upsert_case_file(
                root_module,
                filepath.name,
                "yaml",
                data,
                raw_content=raw,
                source_path=str(filepath.relative_to(ROOT_DIR)),
            )
            migrated_files += 1
    return migrated_files


def main():
    repo.ensure_database()
    env_count = migrate_environments()
    case_file_count = migrate_modules_and_cases()
    print(f"迁移完成: environments={env_count}, case_files={case_file_count}, database=data/tep.db")


if __name__ == "__main__":
    main()
